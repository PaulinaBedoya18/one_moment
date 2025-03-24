import openpyxl
import os
import pandas as pd

ARCHIVO_EXCEL = "tareas.xlsx"

def verificar_archivo():
    """Verifica si el archivo existe, si no, lo crea con estructura inicial."""
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Tareas"
        headers = ["ID", "Nombre", "Descripción", "Fecha Límite", "Estado"]
        sheet.append(headers)
        wb.save(ARCHIVO_EXCEL)

def cargar_tareas():
    """Carga las tareas desde el archivo Excel."""
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    sheet = wb.active
    tareas = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tareas.append(list(row))
    return tareas

def guardar_tareas(tareas):
    """Guarda la lista de tareas en el archivo Excel."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tareas"
    headers = ["ID", "Nombre", "Descripción", "Fecha Límite", "Estado"]
    sheet.append(headers)
    for tarea in tareas:
        sheet.append(tarea)
    wb.save(ARCHIVO_EXCEL)

from datetime import datetime


def agregar_tarea(nombre, descripcion, fecha_limite):
    """Agrega una nueva tarea al archivo Excel."""
    tareas = cargar_tareas()
    nueva_tarea = [len(tareas) + 1, nombre, descripcion, fecha_limite, "Pendiente"]
    tareas.append(nueva_tarea)
    guardar_tareas(tareas)
    print("✅ Tarea agregada con éxito.")

def listar_tareas():
    """Lista todas las tareas."""
    tareas = cargar_tareas()
    if not tareas:
        print("📂 No hay tareas registradas.")
    else:
        print("\n📋 Lista de tareas:")
        for tarea in tareas:
            print(f"{tarea[0]}. {tarea[1]} - {tarea[2]} (Fecha límite: {tarea[3]}) - Estado: {tarea[4]}")

def buscar_tarea(nombre):
    """Busca tareas por nombre."""
    tareas = cargar_tareas()
    resultados = [t for t in tareas if nombre.lower() in t[1].lower()]
    if resultados:
        print("\n🔍 Resultados:")
        for t in resultados:
            print(f"{t[0]}. {t[1]} - {t[2]} (Fecha límite: {t[3]}) - Estado: {t[4]}")
    else:
        print("🚫 No se encontraron tareas con ese nombre.")

        

def completar_tarea(id_tarea):
    """Marca una tarea como completada."""
    tareas = cargar_tareas()
    for tarea in tareas:
        if tarea[0] == id_tarea:
            tarea[4] = "Completada"
            guardar_tareas(tareas)
            print("✅ Tarea marcada como completada.")
            return
    print("🚫 No se encontró la tarea.")

def eliminar_tarea(id_tarea):
    """Elimina una tarea por su ID."""
    tareas = cargar_tareas()
    tareas = [t for t in tareas if t[0] != id_tarea]
    guardar_tareas(tareas)
    print("🗑️ Tarea eliminada.")

def generar_reporte():
    """Genera un reporte de tareas usando pandas."""
    df = pd.read_excel(ARCHIVO_EXCEL)
    print("\n📊 Reporte de tareas:")
    print(df)
    print(f"\nTotal tareas: {len(df)}")
    print(f"Tareas pendientes: {len(df[df['Estado'] == 'Pendiente'])}")
    print(f"Tareas completadas: {len(df[df['Estado'] == 'Completada'])}")

def menu():
    while True:
        print("\n=== GESTOR DE TAREAS ===")
        print("1. Agregar tarea")
        print("2. Listar tareas")
        print("3. Buscar tarea")
        print("4. Completar tarea")
        print("5. Eliminar tarea")
        print("6. Generar reporte")
        print("7. Salir")

        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            nombre = input("Nombre: ")
            descripcion = input("Descripción: ")
            fecha = input("Fecha límite (YYYY-MM-DD): ")
            agregar_tarea(nombre, descripcion, fecha)

        elif opcion == "2":
            listar_tareas()  # ✅ Llamar la función correspondiente

        elif opcion == "3":
            nombre = input("Ingrese el nombre de la tarea a buscar: ")
            buscar_tarea(nombre)  # ✅ Llamar la función correspondiente

        elif opcion == "4":
            id_tarea = int(input("Ingrese el ID de la tarea a completar: "))
            completar_tarea(id_tarea)  # ✅ Llamar la función correspondiente

        elif opcion == "5":
            id_tarea = int(input("Ingrese el ID de la tarea a eliminar: "))
            eliminar_tarea(id_tarea)  # ✅ Llamar la función correspondiente

        elif opcion == "6":
            generar_reporte()  # ✅ Llamar la función correspondiente

        elif opcion == "7":
            print("👋 Saliendo del gestor de tareas...")
            break  # ✅ Sale del bucle

        else:
            print("❌ Opción no válida.")

if __name__ == "__main__":
    verificar_archivo()
    menu()